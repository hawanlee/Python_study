from openpyxl import load_workbook
import datetime
import time
import math
import os

print('''
Author:   LiHaiwang
Date:     2019/06/01
Version:  1.0.0
使用方法：
  1. 将网页上加班时间信息框选对应复制到'./Time_Template.xlsx'中
  2. 运行.exe
  3. 生成的新'./20190601加班时间.xlsx'即含有计算后的加班详细信息
  ''')

wb = load_workbook('./Time_Template.xlsx')
ws = wb['Sheet1']

ws['L1']='间隔'
ws['M1']='起止时间'

# ! 将表格日期转换为datetime标准格式
def convert_time(t):
  month = t.split('/')[0]
  day = t.split('/')[1]
  year = t.split('/')[2][0:4]
  hour = t.split(' ')[-1].split(':')[0]
  minute = t.split(' ')[-1].split(':')[1]
  if t[-2:]=='PM':
    hour = str(int(hour)+12)
  a = datetime.datetime.now()
  return a.replace(int(year), int(month), int(day), int(hour), int(minute), 0, 0)
  
for i in range(1, 50):
  date_item = ws['B'+str(i)].value
  date_type = ws['C'+str(i)].value
  all_time = ws['E'+str(i)].value
  start_time = ws['F'+str(i)].value
  end_time = ws['G'+str(i)].value

  if all_time!=None and start_time!= None and end_time!=None:
    if (date_type=='平时' and all_time>=10.4 ):
      gap_time = (convert_time(end_time)-convert_time(start_time)).total_seconds()/3600-8.5
      fill_start = convert_time(start_time)+datetime.timedelta(hours=8.5)
      fill_end = convert_time(end_time)+datetime.timedelta(hours=0)
      # ws['J'+str(i)]=fill_start.strftime('%H:%M')
      # ws['K'+str(i)]=fill_end.strftime('%H:%M')
      ws['L'+str(i)]=round(all_time - 8.5, 2)
      ws['M'+str(i)]=fill_start.strftime('%H:%M')+' - '+fill_end.strftime('%H:%M')
      ws['N'+str(i)] = convert_time(end_time).strftime(r'%Y/%m/%d')

    if ((date_type=='周末' or date_type=='节假日') and all_time>=2):
      gap_time = (convert_time(end_time)-convert_time(start_time)).total_seconds()/3600
      fill_start = convert_time(start_time)+datetime.timedelta(hours=0)
      fill_end = convert_time(end_time)+datetime.timedelta(hours=0)
      # ws['J'+str(i)]=fill_start.strftime('%H:%M')
      # ws['K'+str(i)]=fill_end.strftime('%H:%M')
      ws['L'+str(i)]=round(all_time, 2)
      ws['M'+str(i)]=fill_start.strftime('%H:%M')+' - '+fill_end.strftime('%H:%M')
      ws['N'+str(i)] = convert_time(end_time).strftime(r'%Y/%m/%d')

e = datetime.datetime.now()
e = str(e.strftime('%Y%m%d'))

time.sleep(1)
print('Working...')
print('')
time.sleep(1)
f = ''+e+'加班填写.xlsx'
wb.save(f)
print(e+'加班填写.xlsx')
print('Opening...')
time.sleep(1)
os.startfile(f)

# 生成.exe
# pyinstaller -F -i test.ico test.py